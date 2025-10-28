
# Create Excel file with both structured and unstructured data sheets
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load the CSV files
df_structured = pd.read_csv('structured_data.csv')
df_unstructured = pd.read_csv('unstructured_data.csv')

# Create Excel file with multiple sheets
output_file = 'excel_ai_data.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
    df_structured.to_excel(writer, sheet_name='Structured_Data', index=False)
    df_unstructured.to_excel(writer, sheet_name='Unstructured_Data', index=False)

print(f"✓ Excel file created: {output_file}")
print(f"  - Sheet 1: Structured_Data ({df_structured.shape[0]} rows, {df_structured.shape[1]} columns)")
print(f"  - Sheet 2: Unstructured_Data ({df_unstructured.shape[0]} rows, {df_unstructured.shape[1]} columns)")

# Format the Excel file
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_file)
print("\n✓ Excel file formatted with headers and column widths")
print(f"\nTotal data generated:")
print(f"  - Structured: {df_structured.shape[0]} rows × {df_structured.shape[1]} columns")
print(f"  - Unstructured: {df_unstructured.shape[0]} rows × {df_unstructured.shape[1]} columns")
