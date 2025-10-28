"""
Excel AI Engine - Main Application (Gemini Version)
Handles file upload, natural language query processing, and data operations
Uses Google Gemini API instead of OpenAI
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, Dict, Any
import pandas as pd
import numpy as np
from datetime import datetime
import os
import json
from pathlib import Path
import logging
from langchain_experimental.agents import create_pandas_dataframe_agent
from langchain_google_genai import ChatGoogleGenerativeAI
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI(
    title="Excel AI Engine (Gemini)",
    description="AI-powered Excel data analysis and manipulation engine using Google Gemini",
    version="1.0.0"
)

# Data models
class AnalysisRequest(BaseModel):
    file_path: str
    query: str
    sheet_name: Optional[str] = None

class AnalysisResponse(BaseModel):
    status: str
    result: Any
    metadata: Optional[Dict] = None

# Global configuration
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# Initialize Gemini LLM
def get_llm():
    """Initialize and return Gemini LLM instance"""
    try:
        api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY or GOOGLE_API_KEY not found in environment")

        llm = ChatGoogleGenerativeAI(
            model="gemini-1.5-flash",  # Free tier model
            google_api_key=api_key,
            temperature=0,
            convert_system_message_to_human=True
        )
        return llm
    except Exception as e:
        logger.error(f"Failed to initialize Gemini LLM: {e}")
        raise HTTPException(status_code=500, detail=f"LLM initialization failed: {str(e)}")

# Excel file operations
class ExcelDataHandler:
    """Handle Excel file reading and writing operations"""

    @staticmethod
    def read_excel(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """Read Excel file and return DataFrame"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            else:
                # Read first sheet if no sheet name provided
                df = pd.read_excel(file_path, engine='openpyxl')
            logger.info(f"Loaded {len(df)} rows from {file_path}")
            return df
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    @staticmethod
    def get_sheet_names(file_path: str) -> list:
        """Get all sheet names from Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            return wb.sheetnames
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    @staticmethod
    def write_excel(df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1"):
        """Write DataFrame to Excel file"""
        try:
            df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
            logger.info(f"Saved {len(df)} rows to {file_path}")
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to write Excel file: {str(e)}")

# Data operations engine
class DataOperationsEngine:
    """Execute various data operations on DataFrame"""

    @staticmethod
    def basic_math(df: pd.DataFrame, operation: str, col1: str, col2: str, result_col: str) -> pd.DataFrame:
        """Perform basic math operations"""
        try:
            if operation == "add":
                df[result_col] = df[col1] + df[col2]
            elif operation == "subtract":
                df[result_col] = df[col1] - df[col2]
            elif operation == "multiply":
                df[result_col] = df[col1] * df[col2]
            elif operation == "divide":
                df[result_col] = df[col1] / df[col2]
            return df
        except Exception as e:
            raise ValueError(f"Math operation failed: {str(e)}")

    @staticmethod
    def aggregations(df: pd.DataFrame, columns: list, agg_funcs: list) -> pd.DataFrame:
        """Calculate aggregations"""
        try:
            result = {}
            for col in columns:
                for func in agg_funcs:
                    if func == "sum":
                        result[f"{col}_sum"] = df[col].sum()
                    elif func == "mean":
                        result[f"{col}_mean"] = df[col].mean()
                    elif func == "min":
                        result[f"{col}_min"] = df[col].min()
                    elif func == "max":
                        result[f"{col}_max"] = df[col].max()
                    elif func == "count":
                        result[f"{col}_count"] = df[col].count()
                    elif func == "std":
                        result[f"{col}_std"] = df[col].std()
            return pd.DataFrame([result])
        except Exception as e:
            raise ValueError(f"Aggregation failed: {str(e)}")

    @staticmethod
    def join_dataframes(df1: pd.DataFrame, df2: pd.DataFrame, 
                       on: str, how: str = "inner") -> pd.DataFrame:
        """Join two DataFrames"""
        try:
            return pd.merge(df1, df2, on=on, how=how)
        except Exception as e:
            raise ValueError(f"Join operation failed: {str(e)}")

    @staticmethod
    def pivot_table(df: pd.DataFrame, values: str, index: str, 
                   columns: str, aggfunc: str = "mean") -> pd.DataFrame:
        """Create pivot table"""
        try:
            return pd.pivot_table(df, values=values, index=index, 
                                 columns=columns, aggfunc=aggfunc)
        except Exception as e:
            raise ValueError(f"Pivot operation failed: {str(e)}")

    @staticmethod
    def unpivot_table(df: pd.DataFrame, id_vars: list, value_vars: list) -> pd.DataFrame:
        """Unpivot (melt) table"""
        try:
            return pd.melt(df, id_vars=id_vars, value_vars=value_vars)
        except Exception as e:
            raise ValueError(f"Unpivot operation failed: {str(e)}")

    @staticmethod
    def date_operations(df: pd.DataFrame, date_col: str) -> pd.DataFrame:
        """Extract date components and perform date operations"""
        try:
            df[date_col] = pd.to_datetime(df[date_col])
            df[f"{date_col}_year"] = df[date_col].dt.year
            df[f"{date_col}_month"] = df[date_col].dt.month
            df[f"{date_col}_day"] = df[date_col].dt.day
            df[f"{date_col}_dayofweek"] = df[date_col].dt.dayofweek
            return df
        except Exception as e:
            raise ValueError(f"Date operation failed: {str(e)}")

    @staticmethod
    def filter_data(df: pd.DataFrame, condition: str) -> pd.DataFrame:
        """Filter data based on condition"""
        try:
            return df.query(condition)
        except Exception as e:
            raise ValueError(f"Filter operation failed: {str(e)}")

# LangChain agent for natural language queries
class NLQueryProcessor:
    """Process natural language queries using LangChain with Gemini"""

    def __init__(self, llm):
        self.llm = llm

    def process_query(self, df: pd.DataFrame, query: str) -> Any:
        """Process natural language query on DataFrame"""
        try:
            # Create pandas DataFrame agent with Gemini
            agent = create_pandas_dataframe_agent(
                self.llm,
                df,
                agent_type="tool-calling",  # Updated for newer LangChain
                verbose=True,
                allow_dangerous_code=True,
                handle_parsing_errors=True
            )

            # Execute query
            result = agent.invoke({"input": query})
            return result
        except Exception as e:
            logger.error(f"Query processing failed: {e}")
            raise ValueError(f"Failed to process query: {str(e)}")

# API Endpoints
@app.get("/")
async def root():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "Excel AI Engine (Gemini)",
        "version": "1.0.0"
    }

@app.get("/health")
async def health_check():
    """Detailed health check"""
    gemini_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "components": {
            "api": "operational",
            "llm": "configured" if gemini_key else "not configured",
            "provider": "Google Gemini"
        }
    }

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload Excel file for processing

    Args:
        file: Excel file (.xlsx)

    Returns:
        File information and sheet names
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

        # Save uploaded file
        file_path = UPLOAD_DIR / file.filename
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Get sheet information
        sheet_names = ExcelDataHandler.get_sheet_names(str(file_path))

        # Read first sheet to get basic info
        df = ExcelDataHandler.read_excel(str(file_path))

        return {
            "status": "success",
            "filename": file.filename,
            "file_path": str(file_path),
            "sheets": sheet_names,
            "rows": len(df),
            "columns": list(df.columns),
            "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()}
        }

    except Exception as e:
        logger.error(f"File upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze")
async def analyze_data(
    file_path: str = Form(...),
    query: str = Form(...),
    sheet_name: Optional[str] = Form(None)
):
    """
    Analyze Excel data using natural language query with Gemini

    Args:
        file_path: Path to uploaded Excel file
        query: Natural language query describing the desired operation
        sheet_name: Optional sheet name to analyze

    Returns:
        Analysis results

    Examples of queries:
    - "Calculate the average salary by department"
    - "Show me the top 10 employees by performance score"
    - "What is the correlation between years of experience and salary?"
    - "Filter employees with salary greater than 100000"
    - "Create a summary showing average salary by department"
    """
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        # Read Excel file
        df = ExcelDataHandler.read_excel(file_path, sheet_name)

        # Initialize Gemini LLM and processor
        llm = get_llm()
        processor = NLQueryProcessor(llm)

        # Process query
        result = processor.process_query(df, query)

        return {
            "status": "success",
            "query": query,
            "result": result,
            "metadata": {
                "rows_analyzed": len(df),
                "columns": list(df.columns),
                "timestamp": datetime.now().isoformat(),
                "llm_provider": "Google Gemini"
            }
        }

    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/operations/math")
async def math_operation(
    file_path: str = Form(...),
    operation: str = Form(...),
    col1: str = Form(...),
    col2: str = Form(...),
    result_col: str = Form(...),
    sheet_name: Optional[str] = Form(None)
):
    """
    Perform basic math operations (add, subtract, multiply, divide)
    """
    try:
        df = ExcelDataHandler.read_excel(file_path, sheet_name)

        operations_engine = DataOperationsEngine()
        df_result = operations_engine.basic_math(df, operation, col1, col2, result_col)

        # Save result
        output_path = OUTPUT_DIR / f"math_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ExcelDataHandler.write_excel(df_result, str(output_path))

        return {
            "status": "success",
            "operation": operation,
            "result_file": str(output_path),
            "sample_results": df_result[[col1, col2, result_col]].head(10).to_dict('records')
        }

    except Exception as e:
        logger.error(f"Math operation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/operations/aggregate")
async def aggregate_data(
    file_path: str = Form(...),
    columns: str = Form(...),
    functions: str = Form(...),
    sheet_name: Optional[str] = Form(None)
):
    """
    Calculate aggregations (sum, mean, min, max, count, std)
    """
    try:
        df = ExcelDataHandler.read_excel(file_path, sheet_name)

        col_list = [c.strip() for c in columns.split(',')]
        func_list = [f.strip() for f in functions.split(',')]

        operations_engine = DataOperationsEngine()
        df_result = operations_engine.aggregations(df, col_list, func_list)

        return {
            "status": "success",
            "aggregations": df_result.to_dict('records')[0]
        }

    except Exception as e:
        logger.error(f"Aggregation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/operations/filter")
async def filter_data(
    file_path: str = Form(...),
    condition: str = Form(...),
    sheet_name: Optional[str] = Form(None)
):
    """
    Filter data based on condition
    """
    try:
        df = ExcelDataHandler.read_excel(file_path, sheet_name)

        operations_engine = DataOperationsEngine()
        df_filtered = operations_engine.filter_data(df, condition)

        # Save result
        output_path = OUTPUT_DIR / f"filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ExcelDataHandler.write_excel(df_filtered, str(output_path))

        return {
            "status": "success",
            "condition": condition,
            "rows_before": len(df),
            "rows_after": len(df_filtered),
            "result_file": str(output_path),
            "sample_results": df_filtered.head(10).to_dict('records')
        }

    except Exception as e:
        logger.error(f"Filter operation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/operations/pivot")
async def create_pivot(
    file_path: str = Form(...),
    values: str = Form(...),
    index: str = Form(...),
    columns: str = Form(...),
    aggfunc: str = Form("mean"),
    sheet_name: Optional[str] = Form(None)
):
    """
    Create pivot table
    """
    try:
        df = ExcelDataHandler.read_excel(file_path, sheet_name)

        operations_engine = DataOperationsEngine()
        df_pivot = operations_engine.pivot_table(df, values, index, columns, aggfunc)

        # Save result
        output_path = OUTPUT_DIR / f"pivot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ExcelDataHandler.write_excel(df_pivot.reset_index(), str(output_path))

        return {
            "status": "success",
            "result_file": str(output_path),
            "pivot_shape": df_pivot.shape,
            "sample_results": df_pivot.head(10).to_dict()
        }

    except Exception as e:
        logger.error(f"Pivot operation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
